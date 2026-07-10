"""
parser.py

Normalizes Zycus project-plan Excel exports into a common schema.

Enhancements:
- variance_days parsing
- row_type classification
- datetime normalization
- data quality metrics generation
"""

from dataclasses import dataclass, field
from pathlib import Path
import re

import pandas as pd
import openpyxl


COMMON_COLUMNS = [
    "task_name",
    "phase_milestone",
    "area",
    "status",
    "pct_complete",
    "schedule_health",
    "at_risk",
    "on_hold",
    "critical",
    "priority",
    "total_float",
    "start_date",
    "end_date",
    "baseline_start",
    "baseline_finish",
    "variance",
    "variance_days",
    "status_comment",
    "owner",
    "existing_rag",
    "row_type"
]


SCHEMA_MAPS = {
    "layout_a": {
        "Task Name": "task_name",
        "Phase/Milestone": "phase_milestone",
        "Area": "area",
        "Status": "status",
        "% Complete": "pct_complete",
        "Schedule Health": "schedule_health",
        "At Risk?": "at_risk",
        "On Hold?": "on_hold",
        "Critical ?": "critical",
        "Priority": "priority",
        "Total Float": "total_float",
        "Start Date": "start_date",
        "End Date": "end_date",
        "Baseline Start": "baseline_start",
        "Baseline Finish": "baseline_finish",
        "Variance": "variance",
        "Status Comment": "status_comment",
        "Owner": "owner",
        "Ancestors": "ancestors"
    },
    "layout_b": {
        "Task Name": "task_name",
        "Phase/Milestone": "phase_milestone",
        "Area": "area",
        "Status": "status",
        "% Complete": "pct_complete",
        "Schedule Health": "schedule_health",
        "At Risk?": "at_risk",
        "On Hold?": "on_hold",
        "Critical ?": "critical",
        "Priority": "priority",
        "Total Float": "total_float",
        "Start Date": "start_date",
        "End Date": "end_date",
        "Baseline Start": "baseline_start",
        "Baseline Finish": "baseline_finish",
        "Variance": "variance",
        "Status Comment": "status_comment",
        "Owner": "owner",
        "RAG": "existing_rag",
        "Ancestors": "ancestors"
    }
}


@dataclass
class ProjectData:
    project_name: str
    source_file: str
    tasks: pd.DataFrame
    comments: pd.DataFrame
    summary: dict

    data_quality_notes: list = field(default_factory=list)
    data_quality_metrics: dict = field(default_factory=dict)


def _clean_unparseable(value):

    if value is None:
        return None

    if isinstance(value, str):
        if value.strip() in ("", "#UNPARSEABLE"):
            return None

    return value


def _parse_variance_days(value):

    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    match = re.search(r"(-?\d+)", str(value))

    if match:
        return float(match.group(1))

    return None


def _classify_row(record):

    ancestors = record.get("ancestors")
    phase_flag = record.get("phase_milestone")
    task_name = str(record.get("task_name", "")).lower()

    if ancestors == 0:
        return "project"

    if phase_flag is True:
        return "milestone"

    if "phase" in task_name:
        return "phase"

    return "task"


def _find_task_sheet(wb):

    for name in wb.sheetnames:

        if name not in ("Comments", "Summary"):
            return name

    raise ValueError("No task sheet found")


def _detect_schema(header_row):

    if "RAG" in header_row:
        return "layout_b"

    return "layout_a"


def _load_summary(wb):

    ws = wb["Summary"]

    summary = {}

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=2
    ):

        key = row[0].value
        val = row[1].value if len(row) > 1 else None

        if key is not None:
            summary[key] = _clean_unparseable(val)

    return summary


def _load_comments(wb):

    if "Comments" not in wb.sheetnames:
        return pd.DataFrame(
            columns=[
                "row_ref",
                "comment_text",
                "author",
                "timestamp"
            ]
        )

    ws = wb["Comments"]

    records = []

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=4,
        values_only=True
    ):

        row_ref, text, author, ts = (
            list(row) + [None, None, None, None]
        )[:4]

        if text:

            records.append(
                {
                    "row_ref": row_ref,
                    "comment_text": text,
                    "author": author,
                    "timestamp": ts
                }
            )

    return pd.DataFrame(
        records,
        columns=[
            "row_ref",
            "comment_text",
            "author",
            "timestamp"
        ]
    )


def load_project(path: str) -> ProjectData:

    path = Path(path)

    wb = openpyxl.load_workbook(
        path,
        data_only=True
    )

    notes = []

    task_sheet_name = _find_task_sheet(wb)
    ws = wb[task_sheet_name]

    header_row = [
        c.value
        for c in next(
            ws.iter_rows(min_row=1, max_row=1)
        )
    ]

    schema_key = _detect_schema(header_row)
    col_map = SCHEMA_MAPS[schema_key]

    header_idx = {
        name: i + 1
        for i, name in enumerate(header_row)
        if name
    }

    missing_cols = [
        c
        for c in col_map
        if c not in header_idx
    ]

    if missing_cols:
        notes.append(
            f"Expected columns missing: {missing_cols}"
        )

    records = []

    for r in range(2, ws.max_row + 1):

        rec = {}

        for src_col, common_col in col_map.items():

            if src_col not in header_idx:

                rec[common_col] = None
                continue

            raw = ws.cell(
                row=r,
                column=header_idx[src_col]
            ).value

            rec[common_col] = _clean_unparseable(raw)

        if rec.get("task_name") is None:
            continue

        rec["variance_days"] = _parse_variance_days(
            rec.get("variance")
        )

        rec["row_type"] = _classify_row(rec)

        records.append(rec)

    tasks = pd.DataFrame(records)

    for col in COMMON_COLUMNS:

        if col not in tasks.columns:
            tasks[col] = None

    date_columns = [
        "start_date",
        "end_date",
        "baseline_start",
        "baseline_finish"
    ]

    for col in date_columns:

        tasks[col] = pd.to_datetime(
            tasks[col],
            errors="coerce"
        )

    project_name = task_sheet_name

    if "ancestors" in tasks.columns:

        root_rows = tasks[
            tasks["ancestors"] == 0
        ]

        if (
            not root_rows.empty
            and root_rows.iloc[0]["task_name"]
        ):

            project_name = root_rows.iloc[0]["task_name"]

    summary = _load_summary(wb)
    comments = _load_comments(wb)

    null_pct_complete = tasks["pct_complete"].isna().mean()

    if null_pct_complete > 0.05:

        notes.append(
            f"{null_pct_complete:.0%} tasks missing % Complete"
        )

    null_schedule_health = (
        tasks["schedule_health"]
        .isna()
        .mean()
    )

    if null_schedule_health > 0.05:

        notes.append(
            f"{null_schedule_health:.0%} tasks missing Schedule Health"
        )

    if not any(
        "budget" in str(h).lower()
        or "cost" in str(h).lower()
        for h in header_row
        if h
    ):
        notes.append(
            "No budget/cost column present in source data"
        )

    data_quality_metrics = {
        "missing_status_pct": round(
            tasks["status"].isna().mean(),
            3
        ),
        "missing_dates_pct": round(
            (
                tasks["start_date"].isna()
                | tasks["end_date"].isna()
            ).mean(),
            3
        ),
        "missing_schedule_health_pct": round(
            tasks["schedule_health"]
            .isna()
            .mean(),
            3
        ),
        "missing_critical_pct": round(
            tasks["critical"]
            .isna()
            .mean(),
            3
        )
    }

    return ProjectData(
        project_name=str(project_name),
        source_file=path.name,
        tasks=tasks,
        comments=comments,
        summary=summary,
        data_quality_notes=notes,
        data_quality_metrics=data_quality_metrics
    )


if __name__ == "__main__":

    for f in [
        "data/Project_Plan_B.xlsx",
        "data/S2P_Project.xlsx",
        "pseudo_data_check/Meridian_Project.xlsx"
    ]:

        d = load_project(f)

        print(f"\n=== {d.project_name} ===")
        print(f"Tasks: {len(d.tasks)}")
        print(f"Comments: {len(d.comments)}")
        print(f"Quality Metrics: {d.data_quality_metrics}")
        print(d.tasks.head(3))